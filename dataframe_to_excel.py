import pandas as pd
from openpyxl import load_workbook

#data =[['Gin',10], ['Henry',15],['Paul',14]]
#df1= pd.DataFrame(data, columns=['Name', 'Age'])

book = load_workbook("C:/Users/henry/Desktop/to be set in proper folder/tetx1.xlsx")
writer = pd.ExcelWriter('C:/Users/henry/Desktop/to be set in proper folder/tetx1.xlsx', engine='openpyxl', mode="a")
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

df1.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)

writer.save()
